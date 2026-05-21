#!/usr/bin/env python3
import argparse
import csv
import datetime as dt
import io
from pathlib import Path

from openpyxl import load_workbook


TARGET_TABLES = [
    "alamat",
    "anak",
    "keluarga",
    "pasangan",
    "pegawai",
    "riwayat_gaji_pokok",
    "riwayat_hukuman_disiplin",
    "riwayat_jabatan",
    "riwayat_keberhasilan",
    "riwayat_kegiatan_strategis",
    "riwayat_narasumber",
    "riwayat_pangkat",
    "riwayat_pendidikan",
    "riwayat_penghargaan",
    "riwayat_prestasi_pendidikan",
    "riwayat_skp",
]

PEGAWAI_COLUMNS = [
    "id_pegawai",
    "nama",
    "jenis_kelamin",
    "tempat_lahir",
    "tanggal_lahir",
    "nik",
    "agama",
    "nama_ukpd",
    "jenis_ukpd",
    "wilayah",
    "jenis_pegawai",
    "status_rumpun",
    "jenis_kontrak",
    "nrk",
    "nip",
    "nama_jabatan_orb",
    "nama_jabatan_menpan",
    "struktur_atasan_langsung",
    "pangkat_golongan",
    "tmt_pangkat_terakhir",
    "jenjang_pendidikan",
    "program_studi",
    "nama_universitas",
    "no_hp_pegawai",
    "email",
    "no_bpjs",
    "kondisi",
    "status_perkawinan",
    "gelar_depan",
    "gelar_belakang",
    "tmt_kerja_ukpd",
    "created_at",
    "id_ukpd",
    "ukpd_id",
    "jenjang_pendidikan_raw",
    "status_rumpun_raw",
    "nama_jabatan_menpan_raw",
    "jenis_kelamin_raw",
]

ALAMAT_COLUMNS = [
    "id",
    "id_pegawai",
    "tipe",
    "jalan",
    "kelurahan",
    "kecamatan",
    "kota_kabupaten",
    "provinsi",
    "kode_provinsi",
    "kode_kota_kab",
    "kode_kecamatan",
    "kode_kelurahan",
    "created_at",
]

KELUARGA_COLUMNS = [
    "id",
    "id_pegawai",
    "hubungan",
    "hubungan_detail",
    "status_punya",
    "status_tunjangan",
    "urutan",
    "nama",
    "jenis_kelamin",
    "tempat_lahir",
    "tanggal_lahir",
    "no_tlp",
    "email",
    "pekerjaan",
    "sumber_tabel",
    "sumber_id",
    "created_at",
]

PASANGAN_COLUMNS = ["id", "id_pegawai", "status_punya", "nama", "no_tlp", "email", "pekerjaan", "created_at"]
ANAK_COLUMNS = ["id", "id_pegawai", "urutan", "nama", "jenis_kelamin", "tempat_lahir", "tanggal_lahir", "pekerjaan", "created_at"]

RIWAYAT_JABATAN_COLUMNS = [
    "id",
    "id_pegawai",
    "nip",
    "nama_pegawai",
    "jenis_jabatan",
    "lokasi",
    "nama_jabatan_orb",
    "nama_jabatan_menpan",
    "struktur_atasan_langsung",
    "nama_ukpd",
    "wilayah",
    "jenis_pegawai",
    "status_rumpun",
    "pangkat_golongan",
    "eselon",
    "tmt_jabatan",
    "nomor_sk",
    "tanggal_sk",
    "keterangan",
    "sumber",
    "source_key",
]

RIWAYAT_PANGKAT_COLUMNS = [
    "id",
    "id_pegawai",
    "nip",
    "nama_pegawai",
    "pangkat_golongan",
    "tmt_pangkat",
    "lokasi",
    "nomor_sk",
    "tanggal_sk",
    "keterangan",
    "sumber",
    "source_key",
]

RIWAYAT_PENDIDIKAN_COLUMNS = [
    "id",
    "id_pegawai",
    "nip",
    "nama_pegawai",
    "jenis_riwayat",
    "jenjang_pendidikan",
    "program_studi",
    "nama_institusi",
    "nama_universitas",
    "kota_institusi",
    "tahun_lulus",
    "nomor_ijazah",
    "tanggal_ijazah",
    "keterangan",
    "sumber",
    "source_key",
]

TABLE_COLUMNS = {
    "pegawai": PEGAWAI_COLUMNS,
    "alamat": ALAMAT_COLUMNS,
    "keluarga": KELUARGA_COLUMNS,
    "pasangan": PASANGAN_COLUMNS,
    "anak": ANAK_COLUMNS,
    "riwayat_jabatan": RIWAYAT_JABATAN_COLUMNS,
    "riwayat_pangkat": RIWAYAT_PANGKAT_COLUMNS,
    "riwayat_pendidikan": RIWAYAT_PENDIDIKAN_COLUMNS,
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    if not text or text in {"-", "#N/A", "NULL", "None"}:
        return None
    if text.endswith(" 00:00:00"):
        return text[:10]
    return text


def normalize_header(value):
    return " ".join(str(value or "").replace("\n", " ").split()).strip()


def normalize_gender(value):
    text = (clean(value) or "").upper().replace(" ", "")
    if text in {"P", "PEREMPUAN", "WANITA"}:
        return "Perempuan"
    if text in {"L", "LAKI-LAKI", "LAKILAKI", "PRIA"}:
        return "Laki-laki"
    return None


def normalize_education(value):
    raw = clean(value)
    if not raw:
        return None
    text = raw.upper().replace(".", " ").replace("_", " ")
    if "S-3" in text or "S 3" in text or "DOKTOR" in text:
        return "S3"
    if "SPESIALIS" in text:
        return "Spesialis"
    if "S-2" in text or "S 2" in text or "MAGISTER" in text:
        return "S2"
    if "PROFESI" in text or "NERS" in text or "APOTEKER" in text or "DOKTER" in text:
        return "Profesi"
    if "S-1" in text or "S 1" in text or "SARJANA" in text:
        return "S1"
    if "D-IV" in text or "D IV" in text or "D4" in text:
        return "D4"
    if "D-III" in text or "D III" in text or "D3" in text:
        return "D3"
    if "SMA" in text or "SMK" in text or "SLTA" in text:
        return "SMA/SMK"
    if "SMP" in text or "SLTP" in text:
        return "SMP"
    if text == "SD" or " SD " in f" {text} ":
        return "SD"
    return raw


def copy_value(value):
    value = clean(value)
    if value is None:
        return r"\N"
    text = str(value)
    return (
        text.replace("\\", "\\\\")
        .replace("\t", r"\t")
        .replace("\n", r"\n")
        .replace("\r", r"\r")
    )


def qident(name):
    return '"' + name.replace('"', '""') + '"'


def write_copy(handle, table, columns, rows):
    if not rows:
        return
    handle.write(f"COPY {qident(table)} ({', '.join(qident(column) for column in columns)}) FROM stdin;\n")
    for row in rows:
        handle.write("\t".join(copy_value(row.get(column)) for column in columns))
        handle.write("\n")
    handle.write("\\.\n\n")


def write_schema_guard(handle):
    handle.write("DO $$\n")
    handle.write("DECLARE\n")
    handle.write("  missing text;\n")
    handle.write("BEGIN\n")
    for table in TARGET_TABLES:
        handle.write(
            "  IF to_regclass('public.{table}') IS NULL THEN\n"
            "    RAISE EXCEPTION 'Required table public.{table} does not exist';\n"
            "  END IF;\n".format(table=table)
        )
    for table, columns in TABLE_COLUMNS.items():
        joined = ", ".join("'" + column.replace("'", "''") + "'" for column in columns)
        handle.write(
            "  SELECT string_agg(column_name, ', ' ORDER BY column_name) INTO missing\n"
            "  FROM (VALUES {values}) AS required(column_name)\n"
            "  WHERE NOT EXISTS (\n"
            "    SELECT 1 FROM information_schema.columns\n"
            "    WHERE table_schema = 'public'\n"
            "      AND table_name = '{table}'\n"
            "      AND column_name = required.column_name\n"
            "  );\n"
            "  IF missing IS NOT NULL THEN\n"
            "    RAISE EXCEPTION 'Table {table} is missing required columns: %', missing;\n"
            "  END IF;\n".format(
                table=table,
                values=", ".join(f"({item})" for item in joined.split(", ")),
            )
        )
    handle.write("END $$;\n\n")


def row_getter(headers, values):
    row = {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}
    return lambda key: clean(row.get(key))


def build_rows(xlsx_path):
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    headers = [normalize_header(value) for value in next(rows_iter)]

    today = dt.date.today().isoformat()
    ukpd_ids = {}
    output = {table: [] for table in TARGET_TABLES}
    counters = {table: 1 for table in TARGET_TABLES}

    def next_id(table):
        value = counters[table]
        counters[table] += 1
        return value

    for index, values in enumerate(rows_iter, start=2):
        get = row_getter(headers, values)
        nama = get("NAMA (TANPA GELAR)")
        nama_ukpd = get("NAMA UKPD")
        if not nama and not nama_ukpd:
            continue

        id_pegawai = int(float(get("No") or len(output["pegawai"]) + 1))
        if nama_ukpd not in ukpd_ids:
            ukpd_ids[nama_ukpd] = len(ukpd_ids) + 1
        id_ukpd = ukpd_ids[nama_ukpd]

        gender_raw = get("JENIS KELAMIN (L/P)")
        pendidikan_raw = get("JENJANG PENDIDIKAN (BERDASARKAN SK PANGKAT TERAKHIR)")
        jabatan_menpan_raw = get("NAMA JABATAN KEPMENPAN 11 TAHUN 2024") or get("NAMA JABATAN PERMENPAN RB NO 41 TAHUN 2018")
        status_rumpun_raw = get("STATUS RUMPUN")

        pegawai = {
            "id_pegawai": id_pegawai,
            "nama": nama,
            "jenis_kelamin": normalize_gender(gender_raw),
            "tempat_lahir": get("TEMPAT LAHIR"),
            "tanggal_lahir": get("TANGGAL LAHIR"),
            "nik": get("NIK"),
            "agama": get("AGAMA"),
            "nama_ukpd": nama_ukpd,
            "jenis_ukpd": get("JENIS UKPD"),
            "wilayah": get("WILAYAH"),
            "jenis_pegawai": get("JENIS PEGAWAI"),
            "status_rumpun": status_rumpun_raw,
            "jenis_kontrak": get("JENIS KONTRAK"),
            "nrk": get("NRK"),
            "nip": get("NIP"),
            "nama_jabatan_orb": get("NAMA JABATAN ORB (PERGUB 1 TAHUN 2017)"),
            "nama_jabatan_menpan": jabatan_menpan_raw,
            "struktur_atasan_langsung": get("STRUKTUR ATASAN LANGSUNG"),
            "pangkat_golongan": get("PANGKAT / GOLONGAN"),
            "tmt_pangkat_terakhir": get("TMT PANGKAT TERAKHIR"),
            "jenjang_pendidikan": normalize_education(pendidikan_raw),
            "program_studi": get("PROGRAM STUDI"),
            "nama_universitas": get("NAMA UNIVERSITAS"),
            "no_hp_pegawai": get("NO. HP PEGAWAI"),
            "email": get("EMAIL AKTIF PEGAWAI"),
            "no_bpjs": get("No BPJS"),
            "kondisi": get("KONDISI"),
            "status_perkawinan": get("Status Perkawinan"),
            "gelar_depan": get("Gelar Depan"),
            "gelar_belakang": get("Gelar Belakang"),
            "tmt_kerja_ukpd": get("TMT KERJA DI UKPD SAAT INI"),
            "created_at": today,
            "id_ukpd": id_ukpd,
            "ukpd_id": id_ukpd,
            "jenjang_pendidikan_raw": pendidikan_raw,
            "status_rumpun_raw": status_rumpun_raw,
            "nama_jabatan_menpan_raw": jabatan_menpan_raw,
            "jenis_kelamin_raw": gender_raw,
        }
        output["pegawai"].append(pegawai)

        for tipe, prefix in [("domisili", "DOMISILI"), ("ktp", "KTP")]:
            alamat = {
                "id": next_id("alamat"),
                "id_pegawai": id_pegawai,
                "tipe": tipe,
                "jalan": get(f"{prefix}_JALAN"),
                "kelurahan": get(f"{prefix}_KELURAHAN"),
                "kecamatan": get(f"{prefix}_KECAMATAN"),
                "kota_kabupaten": get(f"{prefix}_KOTA/KABUPATEN"),
                "provinsi": get(f"{prefix}_PROVINSI"),
                "kode_provinsi": get(f"Kode {prefix.title()} Provinsi"),
                "kode_kota_kab": get(f"Kode {prefix.title()} Kota/Kab"),
                "kode_kecamatan": get(f"Kode {prefix.title()} Kecamatan"),
                "kode_kelurahan": get(f"Kode {prefix.title()} Kelurahan"),
                "created_at": today,
            }
            if any(alamat.get(field) for field in ALAMAT_COLUMNS if field not in {"id", "id_pegawai", "tipe", "created_at"}):
                output["alamat"].append(alamat)

        pasangan_nama = get("NAMA_SUAMI/ISTRI")
        if pasangan_nama or get("NO_TELP_SUAMI/ISTRI") or get("EMAIL_SUAMI/ISTRI") or get("PEKERJAAN"):
            pasangan_id = next_id("pasangan")
            pasangan = {
                "id": pasangan_id,
                "id_pegawai": id_pegawai,
                "status_punya": "Ya" if pasangan_nama else "Tidak",
                "nama": pasangan_nama,
                "no_tlp": get("NO_TELP_SUAMI/ISTRI"),
                "email": get("EMAIL_SUAMI/ISTRI"),
                "pekerjaan": get("PEKERJAAN"),
                "created_at": today,
            }
            output["pasangan"].append(pasangan)
            output["keluarga"].append({
                "id": next_id("keluarga"),
                "id_pegawai": id_pegawai,
                "hubungan": "pasangan",
                "hubungan_detail": "Pasangan",
                "status_punya": pasangan["status_punya"],
                "status_tunjangan": None,
                "urutan": None,
                "nama": pasangan["nama"],
                "jenis_kelamin": None,
                "tempat_lahir": None,
                "tanggal_lahir": None,
                "no_tlp": pasangan["no_tlp"],
                "email": pasangan["email"],
                "pekerjaan": pasangan["pekerjaan"],
                "sumber_tabel": "pasangan",
                "sumber_id": pasangan_id,
                "created_at": today,
            })

        for urutan in [1, 2, 3]:
            suffix = f"KE-{urutan}"
            nama_anak = get(f"NAMA ANAK {suffix}")
            if not nama_anak:
                continue
            tempat_key = "TEMPAT LAHIR 2" if urutan == 2 else f"TEMPAT LAHIR ANAK {suffix}"
            anak_id = next_id("anak")
            anak = {
                "id": anak_id,
                "id_pegawai": id_pegawai,
                "urutan": urutan,
                "nama": nama_anak,
                "jenis_kelamin": normalize_gender(get(f"JENIS KELAMIN ANAK {suffix}")),
                "tempat_lahir": get(tempat_key),
                "tanggal_lahir": get(f"TANGGAL LAHIR ANAK {suffix}"),
                "pekerjaan": get(f"PEKERJAAN ANAK {suffix}"),
                "created_at": today,
            }
            output["anak"].append(anak)
            output["keluarga"].append({
                "id": next_id("keluarga"),
                "id_pegawai": id_pegawai,
                "hubungan": "anak",
                "hubungan_detail": f"Anak {urutan}",
                "status_punya": None,
                "status_tunjangan": None,
                "urutan": urutan,
                "nama": anak["nama"],
                "jenis_kelamin": anak["jenis_kelamin"],
                "tempat_lahir": anak["tempat_lahir"],
                "tanggal_lahir": anak["tanggal_lahir"],
                "no_tlp": None,
                "email": None,
                "pekerjaan": anak["pekerjaan"],
                "sumber_tabel": "anak",
                "sumber_id": anak_id,
                "created_at": today,
            })

        if pegawai["nama_jabatan_orb"] or pegawai["nama_jabatan_menpan"]:
            output["riwayat_jabatan"].append({
                "id": next_id("riwayat_jabatan"),
                "id_pegawai": id_pegawai,
                "nip": pegawai["nip"],
                "nama_pegawai": pegawai["nama"],
                "jenis_jabatan": None,
                "lokasi": pegawai["nama_ukpd"],
                "nama_jabatan_orb": pegawai["nama_jabatan_orb"],
                "nama_jabatan_menpan": pegawai["nama_jabatan_menpan"],
                "struktur_atasan_langsung": pegawai["struktur_atasan_langsung"],
                "nama_ukpd": pegawai["nama_ukpd"],
                "wilayah": pegawai["wilayah"],
                "jenis_pegawai": pegawai["jenis_pegawai"],
                "status_rumpun": pegawai["status_rumpun"],
                "pangkat_golongan": pegawai["pangkat_golongan"],
                "eselon": None,
                "tmt_jabatan": pegawai["tmt_kerja_ukpd"],
                "nomor_sk": None,
                "tanggal_sk": None,
                "keterangan": "Diambil dari master pegawai format awal",
                "sumber": "excel_master_pegawai",
                "source_key": f"excel:riwayat_jabatan:{id_pegawai}",
            })

        if pegawai["pangkat_golongan"] or pegawai["tmt_pangkat_terakhir"]:
            output["riwayat_pangkat"].append({
                "id": next_id("riwayat_pangkat"),
                "id_pegawai": id_pegawai,
                "nip": pegawai["nip"],
                "nama_pegawai": pegawai["nama"],
                "pangkat_golongan": pegawai["pangkat_golongan"],
                "tmt_pangkat": pegawai["tmt_pangkat_terakhir"],
                "lokasi": pegawai["nama_ukpd"],
                "nomor_sk": None,
                "tanggal_sk": None,
                "keterangan": "Diambil dari master pegawai format awal",
                "sumber": "excel_master_pegawai",
                "source_key": f"excel:riwayat_pangkat:{id_pegawai}",
            })

        if pegawai["jenjang_pendidikan"] or pegawai["program_studi"] or pegawai["nama_universitas"]:
            output["riwayat_pendidikan"].append({
                "id": next_id("riwayat_pendidikan"),
                "id_pegawai": id_pegawai,
                "nip": pegawai["nip"],
                "nama_pegawai": pegawai["nama"],
                "jenis_riwayat": "pendidikan_terakhir",
                "jenjang_pendidikan": pegawai["jenjang_pendidikan"],
                "program_studi": pegawai["program_studi"],
                "nama_institusi": pegawai["nama_universitas"],
                "nama_universitas": pegawai["nama_universitas"],
                "kota_institusi": None,
                "tahun_lulus": None,
                "nomor_ijazah": None,
                "tanggal_ijazah": None,
                "keterangan": "Diambil dari master pegawai format awal",
                "sumber": "excel_master_pegawai",
                "source_key": f"excel:riwayat_pendidikan:{id_pegawai}",
            })

    return output


def write_sql(output_path, rows_by_table, source_path):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="\n") as handle:
        handle.write("-- Generated import for SI-SDMK master pegawai.\n")
        handle.write(f"-- Source: {source_path}\n")
        handle.write("-- Review and backup production database before running this file.\n\n")
        handle.write("BEGIN;\n")
        handle.write("SET client_min_messages TO warning;\n\n")
        write_schema_guard(handle)
        handle.write("TRUNCATE TABLE\n  ")
        handle.write(",\n  ".join(qident(table) for table in TARGET_TABLES))
        handle.write("\nRESTART IDENTITY;\n\n")

        for table in TARGET_TABLES:
            columns = TABLE_COLUMNS.get(table)
            if columns:
                write_copy(handle, table, columns, rows_by_table[table])

        sequence_tables = {
            "alamat": "id",
            "anak": "id",
            "keluarga": "id",
            "pasangan": "id",
            "pegawai": "id_pegawai",
            "riwayat_jabatan": "id",
            "riwayat_pangkat": "id",
            "riwayat_pendidikan": "id",
        }
        for table, column in sequence_tables.items():
            handle.write(
                "SELECT setval(pg_get_serial_sequence('{table}', '{column}'), "
                "COALESCE((SELECT MAX({qcolumn}) FROM {qtable}), 1), true) "
                "WHERE pg_get_serial_sequence('{table}', '{column}') IS NOT NULL;\n".format(
                    table=table,
                    column=column,
                    qcolumn=qident(column),
                    qtable=qident(table),
                )
            )
        handle.write("\nCOMMIT;\n")


def write_summary(output_path, rows_by_table):
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["table", "rows"])
        for table in TARGET_TABLES:
            writer.writerow([table, len(rows_by_table.get(table, []))])


def main():
    parser = argparse.ArgumentParser(description="Generate PostgreSQL SQL import from master pegawai XLSX.")
    parser.add_argument("xlsx", nargs="?", default="Analisis_Data_Master_Pegawai_Rapih_Format_Awal.xlsx")
    parser.add_argument("--out", default="generated/master-pegawai-import.sql")
    parser.add_argument("--summary", default="generated/master-pegawai-import-summary.csv")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    out_path = Path(args.out)
    summary_path = Path(args.summary)

    rows_by_table = build_rows(xlsx_path)
    write_sql(out_path, rows_by_table, xlsx_path)
    write_summary(summary_path, rows_by_table)

    print(f"Generated: {out_path}")
    print(f"Summary  : {summary_path}")
    for table in TARGET_TABLES:
        print(f"{table}: {len(rows_by_table.get(table, []))}")


if __name__ == "__main__":
    main()
